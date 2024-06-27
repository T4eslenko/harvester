import asyncio  
import os
import time
import openpyxl
from telethon.tl.functions.channels import InviteToChannelRequest
from telethon.tl.functions.contacts import GetContactsRequest, GetBlockedRequest
from telethon.tl.functions.messages import GetDialogsRequest, ImportChatInviteRequest
from telethon.tl.types import InputChannel, InputPhoneContact, User, Chat, Channel, Message, MessageFwdHeader, MessageMediaDocument, PeerChannel, DocumentAttributeFilename
from telethon.sync import TelegramClient, types
from telethon.errors import SessionPasswordNeededError, PhoneCodeInvalidError, PasswordHashInvalidError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from telethon.errors.rpcerrorlist import PeerFloodError, UserPrivacyRestrictedError
from datetime import datetime
from typing import Optional
import re
from jinja2 import Template
import base64
from io import BytesIO
from PIL import Image
from html import escape
from telethon.sync import TelegramClient
from telethon import functions, types
from telethon.tl.functions.contacts import SearchRequest
from telethon.tl.functions.messages import SearchRequest as MessageSearchRequest
from telethon.tl.types import InputMessagesFilterEmpty
from datetime import datetime
from pytz import timezone
from html import escape
from jinja2 import Environment, FileSystemLoader
from telethon.sync import TelegramClient
from telethon.tl.types import PeerChannel, PeerUser, User, Channel, MessageFwdHeader
import zipfile
import shutil
from dotenv import load_dotenv
from allowed_users import ALLOWED_USERS  # Импортируем словарь из отдельного файла

load_dotenv()
allowed_users = ALLOWED_USERS


# Получаем сообщения пользователей и формируем нумерованный список для выбора диалога для скачивания
async def get_user_dialogs(client):
    user_dialogs = []
    users_list = []
    dialogs = []
    i = 0
    dialogs = await client.get_dialogs()
    if dialogs:
        for dialog in dialogs:
            try:
                if isinstance(dialog.entity, User) and not dialog.entity.bot:
                    try:
                        messages = await client.get_messages(dialog.entity, limit=0)
                        count_messages = messages.total
                    except Exception as e:
                        print(f"Ошибка при получении сообщений для пользователя {dialog.entity.id}: {e}")
                        count_messages = "N/A"  # Значение по умолчанию, если сообщения не удалось получить
    
                    user = dialog.entity
                    username = f'@{user.username}' if user.username else ""
                    first_name = user.first_name if user.first_name else ''
                    last_name = user.last_name if user.last_name else ''
    
                    # Используем чистый текст без ANSI escape-кодов
                    user_dialogs.append(
                        f'👉{i}) {first_name} {last_name} {username} (id: {user.id}) <b>[💬 {count_messages}]</b>')
    
                    users_list.append(dialog.entity.id)
                    i += 1
            except Exception as e:
                print(f"Ошибка при обработке диалога {dialog.id}: {e}")
    return user_dialogs, i, users_list



# Выгрузка самих сообщений
async def get_messages_for_html(client, target_dialog, selection, host_bot_id):
    minsk_timezone = timezone('Europe/Minsk')
    messages = []
    messages_count = 0
    first_message_date = None
    last_message_date = None
    forward_sender = None

    try:
        # Информация об объекте (подсоединен к телеграм-клиенту)
        me = await client.get_me()
        userid_client = me.id
        firstname_client = me.first_name
        username_client = f"@{me.username}" if me.username is not None else ''
        lastname_client = me.last_name if me.last_name is not None else ''
        
    except Exception as e:
        print(f"Ошибка при получении информации о пользователе: {e}")
        return

    try:
        if selection in ['70', '75', '750']:  # если выгрузка из канала
            target_dialog_id = target_dialog.id
            title = target_dialog.title 
            
            selected = 'channel_messages'
            template_file = 'template_groups_messages.html'
        elif selection in ['40', '45', '450']:
            target_dialog_id = target_dialog  # в этом случае target_dialog - это ид пользователя
            title = target_dialog_id
            # Информация о собеседнике
            user = await client.get_entity(target_dialog_id)
            username = f'@{user.username}' if user.username else ''
            first_name = user.first_name if user.first_name else ''
            last_name = user.last_name if user.last_name else ''
            user_id = user.id
            
            selected = 'user_messages'
            template_file = 'template_user_messages.html'
            
    except Exception as e:
        print(f"Ошибка при определении title: {e}")
        return

    try:
        async for message in client.iter_messages(target_dialog_id):
            if selected == 'channel_messages':  # если выгрузка из канала
                try:
                    # target_dialog - это итерация конкретного диалога
                    sender_id = message.sender_id if hasattr(message, 'sender_id') else title
                    username = f"@{message.sender.username}" if hasattr(message.sender, 'username') else ''
                    first_name = message.sender.first_name if hasattr(message.sender, 'first_name') else title
                    last_name = message.sender.last_name if hasattr(message.sender, 'last_name') else ''
                except Exception as e:
                    print(f"Ошибка при получении данных sender_id etc при работе с каналом: {e}")
                    return

            # Определяем дату первого и последнего сообщения
            message_time = message.date.astimezone(minsk_timezone).strftime('%d.%m.%Y %H:%M:%S')
            if first_message_date is None or message.date < first_message_date:
                first_message_date = message.date
            if last_message_date is None or message.date > last_message_date:
                last_message_date = message.date

            if message.sender_id == userid_client:
                sender_info = f"{firstname_client}:"
            else:
                sender_info = f"{first_name}:"

            # Обрабатываем репосты
            forward_text = None
            is_forward = False
            if message.forward:
                is_forward = True
                forward_text = escape(message.text) if message.text else None
                try:
                    forward_sender = await get_forwarded_info(client, message)  # Новая фишка
                except Exception as e:
                    forward_sender = f"Ошибка при получении информации о пересланном сообщении: {e}"

            # Обрабатываем ответы
            reply_text = None
            if message.reply_to_msg_id:
                try:
                    original_message = await client.get_messages(target_dialog_id, ids=message.reply_to_msg_id)
                    if original_message:
                        reply_text = escape(original_message.text) if original_message.text else None
                    else:
                        reply_text = None
                except Exception as e:
                    reply_text = f"Ошибка при получении ответа: {e}"

            # Обрабатываем реакции
            reaction_info = ""
            reactions = message.reactions
            if reactions and reactions.recent_reactions:
                try:
                    if selected == 'channel_messages':
                        reaction_info = ""
                        for reaction in reactions.recent_reactions:
                            if hasattr(reaction.peer_id, 'user_id') and reaction.peer_id.user_id:
                                user_id_react = f"id: {reaction.peer_id.user_id}"
                            else:
                                user_id_react = f"администратор группы: {title}"
                            reaction_info += f"{reaction.reaction.emoticon} ({user_id_react}) "
                        
                        # Убираем последний лишний пробел в конце строки, если он есть
                        reaction_info = reaction_info.strip()
            
                    elif selected == 'user_messages':
                        reaction_info = [" ".join(reaction.reaction.emoticon for reaction in reactions.recent_reactions)]
                    
                except Exception as e:
                    reply_text = f"Ошибка при получении реакции: {e}"

            # Обработка медиафайлов
            media_type = None
            if message.media is not None:
                try:
                    if isinstance(message.media, types.MessageMediaPhoto):
                        if selection in ['45', '450', '75', '750']:
                            # Загрузка фото в формате base64
                            photo_bytes = await client.download_media(message.media.photo, file=BytesIO())
                            if photo_bytes:
                                image = Image.open(photo_bytes)
                                original_size = image.size
                                new_size = (original_size[0] // 2, original_size[1] // 2)
                                image = image.resize(new_size)
                                output = BytesIO()
                                image.save(output, format='JPEG', quality=50)
                                encoded_image = base64.b64encode(output.getvalue()).decode('utf-8')
                                image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                                media_type = f'<img src="{image_data_url}" alt="Photo">'
                            else:
                                media_type = 'Photo'
                        else:
                            media_type = 'Photo'
                    elif isinstance(message.media, types.MessageMediaDocument):
                        for attribute in message.media.document.attributes:
                            if isinstance(attribute, types.DocumentAttributeFilename):
                                document_name = attribute.file_name
                                media_type = f"Document: {document_name}"
                                break
                        if media_type is None:
                            media_type = 'Document (Photo, video, etc)'
                    elif isinstance(message.media, types.MessageMediaWebPage):
                        media_type = 'WebPage'
                    elif isinstance(message.media, types.MessageMediaContact):
                        media_type = 'Contact'
                    elif isinstance(message.media, types.MessageMediaGeo):
                        media_type = 'Geo'
                    elif isinstance(message.media, types.MessageMediaVenue):
                        media_type = 'Venue'
                    elif isinstance(message.media, types.MessageMediaGame):
                        media_type = 'Game'
                    elif isinstance(message.media, types.MessageMediaInvoice):
                        media_type = 'Invoice'
                    elif isinstance(message.media, types.MessageMediaPoll):
                        media_type = 'Poll'
                    elif isinstance(message.media, types.MessageMediaDice):
                        media_type = 'Dice'
                    elif isinstance(message.media, types.MessageMediaPhotoExternal):
                        media_type = 'PhotoExternal'
                    else:
                        media_type = 'Unknown'
                except Exception as e:
                    media_type = f"Ошибка при обработке медиа: {e}"

            messages_count += 1

            messages.append({
                'time': message_time,
                'sender_info': sender_info,
                'reply_text': reply_text,
                'forward_text': forward_text,
                'text': escape(message.text) if message.text else '',
                'reactions': reaction_info,
                'media_type': media_type,
                'sender_id': message.sender_id,
                'is_forward': is_forward,
                'forward_sender': forward_sender
            })
    except Exception as e:
        messages.append({
            'time': '',
            'sender_info': 'Ошибка',
            'reply_text': None,
            'forward_text': None,
            'text': f"Ошибка при получении переписки: {e}",
            'reactions': '',
            'media_type': '',
            'sender_id': None
        })

    try:
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template(template_file)
        html_output = template.render(
            firstname_client=firstname_client,
            first_name=first_name,
            messages=messages,
            userid_client=userid_client,
            title=title,
            first_message_date=first_message_date.astimezone(minsk_timezone).strftime('%d.%m.%Y') if first_message_date else '',
            last_message_date=last_message_date.astimezone(minsk_timezone).strftime('%d.%m.%Y') if last_message_date else '',
            messages_count=messages_count
        )
        save_dir = '/app/files_from_harvester'
        if selected == 'channel_messages':
            def sanitize_filename(filename):
                return re.sub(r'[\\/*?:"<>|]', '', filename)

            clean_group_title = sanitize_filename(title)

            if clean_group_title == title:
                filename = f"{title}_chat_messages.html"
            else:
                filename = f"{clean_group_title}_chat_messages.html"

        elif selected == 'user_messages':
            filename = f"{title}_private_messages.html"
            
        file_path = os.path.join(save_dir, filename)
        
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(html_output)
        print(f"HTML-файл сохранен как '{file_path}'")

        #await send_files_to_bot(bot, admin_chat_ids)

    except Exception as e:
        print(f"Ошибка при сохранении медиафайлов: {e}")

    if selection in ['450', '750']:
        try:
            print()
            print("\033[35mСкачиваю медиа, завари кофе...\033[0m")
            await download_media_files(client, target_dialog_id, host_bot_id)
        except Exception as e:
            print(f"Ошибка при скачивании медиафайлов: {e}")

# Вспомогательная асинхронная функция
async def get_forwarded_info(client, message):
    try:
        # Получение id пересланного пользователя или канала
        fwd_user_id = message.fwd_from.from_id.user_id if isinstance(message.fwd_from, MessageFwdHeader) and hasattr(message.fwd_from.from_id, 'user_id') else None
        fwd_channel_id = message.fwd_from.from_id.channel_id if isinstance(message.fwd_from, MessageFwdHeader) and hasattr(message.fwd_from.from_id, 'channel_id') and isinstance(message.fwd_from.from_id, PeerChannel) else None
        fwd_date = message.fwd_from.date if isinstance(message.fwd_from, MessageFwdHeader) and hasattr(message.fwd_from, 'date') else None

        # Данные о пользователе или канале
        fwd_info = {}

        if fwd_user_id or fwd_channel_id:
            if fwd_user_id:
                fwd_info['Источник'] = "пользователь"
                
                try:
                    # Получение информации о пользователе
                    user = await client.get_entity(PeerUser(fwd_user_id))
                    if isinstance(user, User):
                        # Имя и фамилия
                        if user.first_name or user.last_name:
                            name = " ".join(filter(None, [user.first_name, user.last_name]))
                            fwd_info['Имя и фамилия'] = name
                        
                        # Юзернейм
                        if user.username:
                            fwd_info['Юзернейм'] = f"@{user.username}"
                        
                        # ID пользователя
                        fwd_info['ID'] = fwd_user_id
                except Exception as e:
                    print(f"Ошибка при получении информации о пользователе с ID {fwd_user_id}: {e}")
            else:
                fwd_info['Источник'] = "канал"
                
                try:
                    # Получение информации о канале
                    channel = await client.get_entity(PeerChannel(fwd_channel_id))
                    if isinstance(channel, Channel):
                        # Название канала
                        if channel.title:
                            fwd_info['Название канала'] = channel.title
                        
                        # Ссылка на канал
                        if channel.username:
                            fwd_info['Ссылка на канал'] = f"https://t.me/{channel.username}"
                        
                        # ID канала
                        fwd_info['ID'] = fwd_channel_id
                except Exception as e:
                    print(f"Ошибка при получении информации о канале с ID {fwd_channel_id}: {e}")

        # Дата
        if fwd_date:
            fwd_info['Дата'] = fwd_date.strftime('%d.%m.%Y %H:%M:%S')

        # Формируем строку из непустых значений
        forward_sender = ", ".join([f"{key}: {value}" for key, value in fwd_info.items()]) if fwd_info else "Источник неизвестен"
        return forward_sender
    
    except Exception as e:
        print(f"Общая ошибка при обработке пересланного сообщения: {e}")
        return "Источник неизвестен"


#Вспомогательная функция по скачиванию медиа
async def download_media_files(client, target_user, host_bot_id):
    try:
        host_bot_id_str = str(host_bot_id)
        target_user_str = str(target_user)
        nickname = ALLOWED_USERS.get(host_bot_id, "Unknown_User")
        
        # Удаление недопустимых символов из названия
        nickname_clean = re.sub(r'[\\/*?:"<>|]', '', nickname)
        
        # Определение имени файла в зависимости от того, изменилось ли название
        if nickname_clean == nickname:
            user_nickname = nickname
        else:
            user_nickname = nickname_clean

        # Формируем название подпапки на основе user_id, target_user и текущей даты и времени
        current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        user_folder = os.path.join('/app/files_from_harvester', f"{host_bot_id_str}-{user_nickname} выгрузил {target_user_str}-{current_time}")
        os.makedirs(user_folder, exist_ok=True)

        async for message in client.iter_messages(target_user):
            if message.media is not None:
                if isinstance(message.media, (types.MessageMediaPhoto, types.MessageMediaDocument)):
                    try:
                        # Скачивание медиафайла в подпапку пользователя
                        media_path = await client.download_media(message.media, file=user_folder)

                        if media_path:
                            print(f"Скачан медиафайл: {media_path}")

                    except Exception as e:
                        print(f"Ошибка при скачивании медиафайла: {e}")

        print(f"Скачивание медиафайлов для {target_user} завершено")

    except Exception as e:
        print(f"Ошибка при получении сообщений: {e}")








# Получение информации о пользователе
async def get_bot_from_search(client, phone_number, selection, list_botblocked, list_botexisted):
    bot_from_search = []
    bot_from_search_html = []
    try:
        keyword = 'bot'
        entities = await client(SearchRequest(
            q=keyword,
            limit=1000  # Максимальное количество сущностей, которые нужно получить
        ))
        for user in entities.users:
            if user.id not in list_botblocked and user.id not in list_botexisted:

                if user.photo:
                    user_info = await client.get_entity(user.id)
                    #if user_info.photo:
                    photo_path = await client.download_profile_photo(user, file=BytesIO())
                    if photo_path:
                            encoded_image = base64.b64encode(photo_path.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                    else:
                            with open("no_image.png", "rb") as img_file:
                                img_data = img_file.read()
                                img_str = base64.b64encode(img_data).decode('utf-8')
                                image_data_url = f"data:image/png;base64,{img_str}"
                    bot_from_search_html.append(
                            f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                            f'<a href="https://t.me/{user.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{user.username}</a> '
                            f'<span style="color:#556B2F;vertical-align:middle;">{user.first_name}</span>'
                    )
                        
                    bot_from_search.append(f"{user.first_name}, @{user.username}")
    except Exception as e:
        print(f"An error occurred: {e}")

    return bot_from_search, bot_from_search_html
    
async def get_user_info(client, phone, selection):
    """Функция для получения информации о пользователе и его ID."""
    me = await client.get_me()
    userid = me.id
    firstname = me.first_name
    username = f"@{me.username}" if me.username is not None else ""
    lastname = me.last_name if me.last_name is not None else ""
    userinfo = f"(Номер телефона: +{phone}, ID: {userid}, ({firstname} {lastname}) {username})"
    photos_user_html = ''
    if selection == '0':
        try:
            user_photo = await client.get_profile_photos(userid)
            if user_photo:
                for i in range(len(user_photo)):
                    file_name = f"{phone}_{i}"
                    await client.download_media(user_photo[i], file=file_name)
                    jpg_path = f"{file_name}.jpg"
                    mp4_path = f"{file_name}.mp4"
                    if os.path.exists(jpg_path):
                        with open(jpg_path, "rb") as img_file:
                            img_data = open(jpg_path, "rb").read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            photos_user_html += f'<img src="data:image/jpeg;base64,{img_str}" alt="User photo {i+1}" style="width:100px;height:100px;vertical-align:middle;margin-right:10px;">'
                        os.remove(jpg_path)
                    elif os.path.exists(mp4_path):
                        with open(mp4_path, "rb") as video_file:
                            video_data = video_file.read()
                            video_str = base64.b64encode(video_data).decode('utf-8')
                            photos_user_html += f'<video width="100" height="100" controls><source src="data:video/mp4;base64,{video_str}" type="video/mp4">Your browser does not support the video tag.</video>'
                        os.remove(mp4_path)
            else:
                with open("no_image.png", "rb") as img_file:
                    img_data = img_file.read()
                    img_str = base64.b64encode(img_data).decode('utf-8')
                    image_data_url = f"data:image/png;base64,{img_str}"
        
                    photos_user_html +=f'<img src="data:image/png;base64,{img_str}" alt=" " style="width:100px;height:100px;vertical-align:middle;margin-right:10px;">'
        except Exception as e:
            print(f"An error occurred: {e}")
    return userid, userinfo, firstname, lastname, username, photos_user_html



async def get_type_of_chats(client, selection):
    """Функция для подсчета количества сообщений в чатах и определения типов чатов."""
    chat_message_counts = {}
    openchannels = []
    closechannels = []
    openchats = []
    closechats = []
    count_messages = 0
    deactivated_chats = []
    all_chats_ids = []
    delgroups = []
    chats = []
    admin_id = [] 
    user_bots = []
    user_bots_html = []
    image_data_url = ''
    list_botexisted =[]
    chats = []

    
    chats = await client.get_dialogs()
    if chats:
        for chat in chats:   
            # Получаем данные о ботах
            if isinstance(chat.entity, User) and chat.entity.bot: 
                if selection == '0':
                    try:
                        photo_bytes = await client.download_profile_photo(chat.entity, file=BytesIO())
                        if photo_bytes:
                            encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                        else:
                            with open("no_image.png", "rb") as img_file:
                                img_data = img_file.read()
                                img_str = base64.b64encode(img_data).decode('utf-8')
                                image_data_url = f"data:image/png;base64,{img_str}"
                    except Exception:
                        pass
                user_bots_html.append(
                    f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                    f'<a href="https://t.me/{chat.entity.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{chat.entity.username}</a> '
                    f'<span style="color:#556B2F;vertical-align:middle;">{chat.entity.first_name}</span>'
                )
                
                user_bots.append(f"{chat.entity.first_name}, @{chat.entity.username}")
                list_botexisted.append(chat.entity.id)
    
            # Работаем с групповыми чатами
            if isinstance(chat.entity, Channel) or isinstance(chat.entity, Chat):  
                # Выгружаем количество сообщений при выборе опции выгрузить сообщение
                if selection in ['7', '70', '75', '750']: 
                    messages = await client.get_messages(chat.entity, limit=0)
                    count_messages = messages.total
                    chat_message_counts[chat.entity.id] = count_messages
    
                # Определяем открытый канал
                if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast') and chat.entity.participants_count is not None:
                    if chat.entity.broadcast and chat.entity.username:
                        if selection == '6':
                            if chat.entity.admin_rights or chat.entity.creator:
                                openchannels.append(chat.entity)
                                all_chats_ids.append(chat.entity.id)
                                admin_id.append(chat.entity.id)
                        
                        if selection != '6':
                            openchannels.append(chat.entity)
                            all_chats_ids.append(chat.entity.id)
                            if chat.entity.admin_rights or chat.entity.creator:
                                admin_id.append(chat.entity.id)
    
                # Определяем закрытый канал
                if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                    if chat.entity.broadcast and chat.entity.username is None and chat.entity.title != 'Unsupported Chat':
                        if selection == '6':
                            if chat.entity.admin_rights or chat.entity.creator:
                                closechannels.append(chat.entity)
                                all_chats_ids.append(chat.entity.id)
                                admin_id.append(chat.entity.id)
                        
                        if selection != '6':
                            closechannels.append(chat.entity)
                            all_chats_ids.append(chat.entity.id)
                            if chat.entity.admin_rights or chat.entity.creator:
                                admin_id.append(chat.entity.id)
    
                # Определяем открытый чат
                if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                    if not chat.entity.broadcast and chat.entity.username:
                        openchats.append(chat.entity)
                        all_chats_ids.append(chat.entity.id)
                        admin_id.append(chat.entity.id)
    
                # Определяем закрытый чат
                if isinstance(chat.entity, Channel) and hasattr(chat.entity, 'broadcast'):
                   if chat.entity.broadcast == False and chat.entity.username == None:
                      closechats.append(chat.entity)
                      all_chats_ids.append(chat.entity.id)
                      admin_id.append(chat.entity.id)
    
                if isinstance(chat.entity, Chat) and chat.entity.migrated_to is None:
                   closechats.append(chat.entity)
                   all_chats_ids.append(chat.entity.id)
                   admin_id.append(chat.entity.id)
    
                    
                if selection == '5' or selection == '0': # Добавляем нулевые чаты только для общей информации
                    if isinstance(chat.entity, Chat) and hasattr(chat.entity, 'participants_count') and chat.entity.participants_count == 0:
                       if chat.entity.migrated_to is not None and isinstance(chat.entity.migrated_to, InputChannel):
                          deactivated_chats_all = {
                             'ID_migrated': chat.entity.migrated_to.channel_id,
                             'ID': chat.entity.id,
                             'title': chat.entity.title,
                             'creator': chat.entity.creator,
                             'admin_rights': chat.entity.admin_rights,
                          }
                          deactivated_chats.append(deactivated_chats_all)
       
        if selection == '5' or selection == '0': # Добавляем нулевые чаты для общей информации
           if isinstance(chat.entity, Channel) or isinstance(chat.entity, Chat): # Проверяем, является ли чат групповым
              for current_deleted_chat in deactivated_chats:
                     ID_migrated_values = current_deleted_chat['ID_migrated']
                     if ID_migrated_values not in all_chats_ids:
                          delgroups.append(current_deleted_chat)


    return delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, admin_id, user_bots, user_bots_html, list_botexisted



async def get_blocked_bot(client, selection):
    blocked_bot_info = []
    blocked_bot_info_html = []
    count_blocked_bot = 0
    earliest_date = None
    latest_date = None
    image_data_url = " "
    list_botblocked =[]
    
    delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, admin_id, user_bots, user_bots_html, list_botexisted = await get_type_of_chats(client, selection)
    result_blocked = await client(GetBlockedRequest(offset=0, limit=200))
    for peer in result_blocked.blocked:
        if peer.peer_id.__class__.__name__ == 'PeerUser':
            user = await client.get_entity(peer.peer_id.user_id)
            if user.bot:
                if selection == '0':
                    try:
                        photo_path = await client.download_profile_photo(user, file=BytesIO())
                        if photo_path:
                            encoded_image = base64.b64encode(photo_path.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                        else:
                            with open("no_image.png", "rb") as img_file:
                                img_data = img_file.read()
                                img_str = base64.b64encode(img_data).decode('utf-8')
                                image_data_url = f"data:image/png;base64,{img_str}"
                    except Exception:
                        pass    
                blocked_bot_info.append(f"\033[36m@{user.username}\033[0m \033[93m'{user.first_name}'\033[0m заблокирован: {peer.date.strftime('%d/%m/%Y')}")
                
                blocked_bot_info_html.append(
                    f'<img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                    f'<a href="https://t.me/{user.username}" style="color:#0000FF; text-decoration: none;vertical-align:middle;">@{user.username}</a> '
                    f'<span style="color:#556B2F;vertical-align:middle;">{user.first_name}</span> заблокирован: {peer.date.strftime("%d/%m/%Y")}'
                )

                list_botblocked.append(peer.peer_id.user_id)

    return count_blocked_bot, earliest_date, latest_date, blocked_bot_info, blocked_bot_info_html, user_bots, user_bots_html, list_botblocked

# Функция для получения списка прав администратора в канале
def get_admin_rights_channel_list(admin_rights):
    rights = ['<span style="color:maroon; font-weight:bold; font-style:italic;">Права, как администратора канала:</span>']
    possible_rights = {
        'Изменение профиля канала': admin_rights.change_info if admin_rights else False,
        'Публикация сообщений': admin_rights.post_messages if admin_rights else False,
        'Изменение публикаций': admin_rights.edit_messages if admin_rights else False,
        'Удаление публикаций': admin_rights.delete_messages if admin_rights else False,
        'Публикация историй': admin_rights.post_stories if admin_rights else False,
        'Изменение историй': admin_rights.edit_stories if admin_rights else False,
        'Удаление историй': admin_rights.delete_stories if admin_rights else False,
        'Пригласительные ссылки': admin_rights.invite_users if admin_rights else False,
        'Управление трансляциями': admin_rights.manage_call if admin_rights else False,
        '<b>Назначение администраторов</b>': admin_rights.add_admins if admin_rights else False,
    }
    has_any_rights = any(possible_rights.values())
    for right, has_right in possible_rights.items():
        status = '<b><span style="color:red; font-weight:bold;">да</span></b>' if has_right else '<b>нет</b>'
        rights.append(f"{right} - {status}")
    return rights if has_any_rights else []

def get_admin_rights_chat_list(admin_rights):
    rights = ['<span style="color:maroon; font-weight:bold; font-style:italic;">Права, как администратора группы:</span>']
    possible_rights = {
        'Удаление сообщений': admin_rights.delete_messages if admin_rights else False,
        'Блокировка пользователей': admin_rights.ban_users if admin_rights else False,
        'Пригласительные ссылки': admin_rights.invite_users if admin_rights else False,
        'Закрепление сообщений': admin_rights.pin_messages if admin_rights else False,
        'Публикация историй': admin_rights.post_stories if admin_rights else False,
        'Изменение историй': admin_rights.edit_stories if admin_rights else False,
        'Удаление историй': admin_rights.delete_stories if admin_rights else False,
        'Управление трансляциями': admin_rights.manage_call if admin_rights else False,
        '<b>Назначение администраторов</b>': admin_rights.add_admins if admin_rights else False,
        'Анонимность': admin_rights.anonymous if admin_rights else False
    }
    has_any_rights = any(possible_rights.values())
    for right, has_right in possible_rights.items():
        status = '<b><span style="color:red; font-weight:bold;">да</span></b>' if has_right else '<b>нет</b>'
        rights.append(f"{right} - {status}")
    return rights if has_any_rights else []

async def make_list_of_channels(delgroups, chat_message_counts, openchannels, closechannels, openchats, closechats, selection, client):
    """Функция для формирования списков групп и каналов"""
    owner_openchannel = 0
    owner_opengroup = 0
    owner_closegroup = 0
    owner_closechannel = 0
    all_info = []
    groups = []
    i=0
    channels_list = []
    if selection in ['40', '45', '450', '70', '75', '750']:
        openchannel_count =0
        opengroup_count =0
        closegroupdel_count=0
        public_channels_html=[]
        public_groups_html=[]
        deleted_groups_html=[]
        closechannel = []
        closechat = []

    if selection not in ['70', '75', '750']:
        openchannels_name = 'Открытые КАНАЛЫ:' if openchannels else ''
        all_info.append(f"\033[95m{openchannels_name}\033[0m")  
        openchannel_count = 1  
        public_channels_html = []
        image_data_url = ''
        for openchannel in openchannels:
            try:
                photo_bytes = await client.download_profile_photo(openchannel, file=BytesIO())
                if photo_bytes:
                    encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                    image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                    with open("no_image.png", "rb") as img_file:
                        img_data = img_file.read()
                        img_str = base64.b64encode(img_data).decode('utf-8')
                        image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
            count_row = openchannel_count if selection == '5' or selection == '0' else i
            owner = " (Владелец)" if openchannel.creator else ""
            admin = " (Администратор)" if openchannel.admin_rights is not None else ""
            
            # Получение списка прав администратора
            admin_rights_list = get_admin_rights_channel_list(openchannel.admin_rights)
            admin_rights_html = ""
            if admin_rights_list:
                admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
            
            messages_count = f" / [{chat_message_counts.get(openchannel.id, 0)}]" if chat_message_counts else ""
            all_info.append(f"{count_row} - {openchannel.title} \033[93m[{openchannel.participants_count}]{messages_count}\033[0m\033[91m {owner} {admin}\033[0m ID:{openchannel.id} \033[94m@{openchannel.username}\033[0m")
            
            public_channels_html.append(
                f"{openchannel_count}. <img src='{image_data_url}' alt=' ' style='width:50px;height:50px;vertical-align:middle;margin-right:10px;'>"
                f"<span style='color:#556B2F;'>{openchannel.title}</span> <span style='color:#8B4513;'>[{openchannel.participants_count}]</span> "
                f"<span style='color:#FF0000;'>{owner} {admin}</span> ID:{openchannel.id} "
                f'<a href="https://t.me/{openchannel.username}" style="color:#0000FF; text-decoration: none;">@{openchannel.username}</a>'
                f"{admin_rights_html}"
            )
            
            
            openchannel_count += 1
            groups.append(openchannel)
            i += 1
            if owner != "" or admin != "":
                owner_openchannel += 1

    

    closechannels_name = 'Закрытые КАНАЛЫ:' if closechannels else ''
    all_info.append(f"\033[95m{closechannels_name}\033[0m")  
    if closechannels_name:
        channels_list.append(f"💥💥💥 {closechannels_name} 💥💥💥")
    closechannel_count = 1
    private_channels_html = []
    image_data_url = ''
    for closechannel in closechannels:
        if selection == '0':
            try:
                photo_bytes = await client.download_profile_photo(closechannel, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                        with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = closechannel_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if closechannel.creator else ""
        admin = " (Администратор)" if closechannel.admin_rights is not None else ""

        # Получение списка прав администратора
        admin_rights_list = get_admin_rights_channel_list(closechannel.admin_rights)
        admin_rights_html = ""
        if admin_rights_list:
            admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
        
        messages_count = f" / [{chat_message_counts.get(closechannel.id, 0)}]" if chat_message_counts else ""
        messages_count_for_harvester = chat_message_counts.get(closechannel.id, 0) if chat_message_counts else ""
        all_info.append(f"{count_row} - {closechannel.title} \033[93m[{closechannel.participants_count}]{messages_count}\033[0m \033[91m{owner} {admin}\033[0m ID:{closechannel.id}")
        private_channels_html.append(
            f'{closechannel_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{closechannel.title}</span> <span style='color:#8B4513;'>[{closechannel.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechannel.id}"
            f"{admin_rights_html}"
        )

        # Используем чистый текст без ANSI escape-кодов
        channels_list.append(f'👉{i}) {closechannel.title} [🧍 {closechannel.participants_count}, 💬 {messages_count_for_harvester}]')        
        closechannel_count += 1
        groups.append(closechannel)
        i +=1
        if owner != "" or admin != "":
            owner_closechannel += 1

    if selection not in ['70', '75', '750']:

        openchats_name = 'Открытые ГРУППЫ:' if openchats else ''
        all_info.append(f"\033[95m{openchats_name}\033[0m")
        opengroup_count = 1
        public_groups_html = []
        image_data_url = ''
        for openchat in openchats:
            if selection == '0':
                try:
                    photo_bytes = await client.download_profile_photo(openchat, file=BytesIO())
                    if photo_bytes:
                            encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                            image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                    else:
                            with open("no_image.png", "rb") as img_file:
                                img_data = img_file.read()
                                img_str = base64.b64encode(img_data).decode('utf-8')
                                image_data_url = f"data:image/png;base64,{img_str}"
                except Exception:
                    pass 
            count_row = opengroup_count if selection == '5' or selection == '0' else i
            owner = " (Владелец)" if openchat.creator else ""
            admin = " (Администратор)" if openchat.admin_rights is not None else ""
            admin_rights_list = get_admin_rights_chat_list(openchat.admin_rights)
            admin_rights_html = ""
            if admin_rights_list:
                admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
            
            messages_count = f" / [{chat_message_counts.get(openchat.id, 0)}]" if chat_message_counts else ""
            all_info.append(f"{count_row} - {openchat.title} \033[93m[{openchat.participants_count}]{messages_count}\033[0m\033[91m {owner} {admin}\033[0m ID:{openchat.id} \033[94m@{openchat.username}\033[0m")
            public_groups_html.append(
                f'{opengroup_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
                f"<span style='color:#556B2F;'>{openchat.title}</span> <span style='color:#8B4513;'>[{openchat.participants_count}]</span> "
                f"<span style='color:#FF0000;'>{owner} {admin}</span> ID:{openchat.id} "
                f'<a href="https://t.me/{openchat.username}" style="color:#0000FF; text-decoration: none;">@{openchat.username}</a>'
                f"{admin_rights_html}"
            )
            opengroup_count += 1
            groups.append(openchat)
            i +=1
            if owner != "" or admin != "":
                owner_opengroup += 1

    closechats_name = 'Закрытые ГРУППЫ:' if closechats else ''
    if closechats_name:
        channels_list.append(f"\n\n💥💥💥 {closechats_name} 💥💥💥")
    all_info.append(f"\033[95m{closechats_name}\033[0m")
    closegroup_count = 1
    private_groups_html = []
    image_data_url = ''
    for closechat in closechats:
        if selection == '0':
            try:
                photo_bytes = await client.download_profile_photo(closechat, file=BytesIO())
                if photo_bytes:
                        encoded_image = base64.b64encode(photo_bytes.getvalue()).decode('utf-8')
                        image_data_url = f"data:image/jpeg;base64,{encoded_image}"
                else:
                        with open("no_image.png", "rb") as img_file:
                            img_data = img_file.read()
                            img_str = base64.b64encode(img_data).decode('utf-8')
                            image_data_url = f"data:image/png;base64,{img_str}"
            except Exception:
                pass 
        count_row = closegroup_count if selection == '5' or selection == '0' else i
        owner = " (Владелец)" if closechat.creator else ""
        admin = " (Администратор)" if closechat.admin_rights is not None else ""
        admin_rights_list = get_admin_rights_chat_list(closechat.admin_rights)
        admin_rights_html = ""
        if admin_rights_list:
            admin_rights_html = "<ul style='font-size:14px; font-style:italic;'>" + "".join([f"<li style='margin-left:50px;'>{right}</li>" for right in admin_rights_list]) + "</ul>"
        
        messages_count = f" / [{chat_message_counts.get(closechat.id, 0)}]" if chat_message_counts else ""
        messages_count_for_harvester = chat_message_counts.get(closechat.id, 0) if chat_message_counts else ""
        all_info.append(f"{count_row} - {closechat.title} \033[93m[{closechat.participants_count}]{messages_count}\033[0m \033[91m{owner} {admin}\033[0m ID:{closechat.id}")
        private_groups_html.append(
            f'{closegroup_count}. <img src="{image_data_url}" alt=" " style="width:50px;height:50px;vertical-align:middle;margin-right:10px;">'
            f"<span style='color:#556B2F;'>{closechat.title}</span> <span style='color:#8B4513;'>[{closechat.participants_count}]</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{closechat.id}"
            f"{admin_rights_html}"
        )
        channels_list.append(f'👉{i}) {closechat.title} [🧍 {closechat.participants_count}, 💬 {messages_count_for_harvester}]')     
        closegroup_count += 1
        groups.append(closechat)
        i +=1
        if owner != "" or admin != "":
            owner_closegroup += 1

    if selection not in ['70', '75', '750']:
        delgroups_name = 'Удаленные ГРУППЫ:' if delgroups else ''
        all_info.append(f"\033[95m{delgroups_name}\033[0m")
        closegroupdel_count = 1
        deleted_groups_html = []
        for delgroup in delgroups:
            count_row = closegroupdel_count if selection == '5' or selection == '0' else i
            owner_value = delgroup['creator']
            admin_value = delgroup['admin_rights']
            id_value = delgroup['ID']
            title_value = delgroup['title']
            owner = " (Владелец)" if owner_value else ""
            admin = " (Администратор)" if admin_value is not None else ""
            all_info.append(f"{count_row} - {title_value} \033[91m{owner} {admin}\033[0m ID:{id_value}")
            deleted_groups_html.append(f"{closegroupdel_count} - <span style='color:#556B2F;'>{title_value}</span> <span style='color:#FF0000;'>{owner} {admin}</span> ID:{id_value}")
            closegroupdel_count += 1
            i +=1
            if owner != "" or admin != "":
                owner_closegroup += 1

    return groups, i, all_info, openchannel_count, closechannel_count, opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html, private_channels_html, public_groups_html, private_groups_html, deleted_groups_html, channels_list


async def get_and_save_contacts(client, phone_user, userid_user, userinfo, firstname_user, lastname_user, username_user, needsavecontacts):
    result = await client(GetContactsRequest(0))
    contacts = result.users
    total_contacts = len(contacts)
    total_contacts_with_phone = sum(bool(getattr(contact, 'phone', None)) for contact in contacts)
    total_mutual_contacts = sum(bool(getattr(contact, 'mutual_contact', None)) for contact in contacts)

    if needsavecontacts == '1':
        # Сохраняем информацию о контактах
        new_phone_user = phone_user[1:]  # "1234567890"
        contacts_file_name = f'/app/files_from_harvester/{phone_user}_contacts.xlsx'
        print(f"Контакты сохранены в файл {phone_user}_contacts.xlsx")
    
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = ['ID контакта', 'First name контакта', 'Last name контакта', 'Username контакта', 'Телефон контакта', 'Взаимный контакт', 'Дата внесения в базу', 'First name объекта', 'Last name объекта', 'Username объекта', 'Телефон объекта', 'ID_объекта']
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
            
        row_num = 2
        for contact in contacts:
            if hasattr(contact, 'id'):
                sheet.cell(row=row_num, column=1, value=contact.id)
            if hasattr(contact, 'first_name'):
                sheet.cell(row=row_num, column=2, value=contact.first_name)
            if hasattr(contact, 'last_name'):
                sheet.cell(row=row_num, column=3, value=contact.last_name)
            if hasattr(contact, 'username') and contact.username is not None:
                username_with_at = f"@{contact.username}"
                sheet.cell(row=row_num, column=4, value=username_with_at)
            if hasattr(contact, 'phone'):
                sheet.cell(row=row_num, column=5, value=contact.phone)
            if hasattr(contact, 'mutual_contact') and contact.mutual_contact:
                sheet.cell(row=row_num, column=6, value='взаимный')
            sheet.cell(row=row_num, column=7, value=datetime.now().strftime('%d.%m.%Y %H:%M:%S'))
            sheet.cell(row=row_num, column=8, value=firstname_user)
            sheet.cell(row=row_num, column=9, value=lastname_user)
            sheet.cell(row=row_num, column=10, value=username_user)
            sheet.cell(row=row_num, column=11, value=new_phone_user)
            sheet.cell(row=row_num, column=12, value=userid_user)
         
            row_num += 1
    
        wb.save(contacts_file_name)
    return total_contacts, total_contacts_with_phone, total_mutual_contacts

async def save_about_channels(phone, userid, firstname, lastname, username, openchannel_count, opengroup_count, closechannel_count, closegroup_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, openchannels, closechannels, openchats, closechats, delgroups, closegroupdel_count):
    
    async def write_data(sheet, data):
        sheet.append(["Название", "Количество участников", "Владелец", "Администратор", "ID", "Ссылка"])
        for item in data:
          owner = " (Владелец)" if item.creator else ""
          admin = " (Администратор)" if item.admin_rights is not None else ""
          usernameadd = f"@{item.username}" if hasattr(item, 'username') and item.username is not None else ""
          sheet.append([item.title, item.participants_count, owner, admin, item.id, usernameadd])
    
    async def write_data_del(sheet, data):
        sheet.append(["Название", "Владелец", "Администратор", "ID"])
        for item in data:
          owner_value = item['creator']
          admin_value = item['admin_rights']
          id_value = item['ID']
          title_value = item['title']
          owner = " (Владелец)" if owner_value else ""
          admin = " (Администратор)" if admin_value is not None else ""
          sheet.append([title_value, owner, admin, id_value])
            
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws_summury = wb.create_sheet("Сводная информация")
    ws_summury.append([f"Номер телефона: +{phone}, ID: {userid}, ({firstname}{lastname}) {username}"])
    if openchannel_count > 1:
        ws_summury.append([f"Открытые каналы: {openchannel_count-1}"])
        ws_open_channels = wb.create_sheet("Открытые каналы")
        await write_data(ws_open_channels, openchannels)
    if closechannel_count > 1:
        ws_summury.append([f"Закрытые каналы: {closechannel_count-1}"])
        ws_closed_channels = wb.create_sheet("Закрытые каналы")
        await write_data(ws_closed_channels, closechannels)
    if owner_openchannel > 1:
        ws_summury.append([f"Имеет права владельца или админа в открытых каналах: {owner_openchannel}"])
    if owner_closechannel > 1:
        ws_summury.append([f"Имеет права владельца или админа в закрытых каналах: {owner_closechannel}"])
    if opengroup_count > 1:
        ws_summury.append([f"Открытые группы: {opengroup_count-1}"])
        ws_open_groups = wb.create_sheet("Открытые группы")
        await write_data(ws_open_groups, openchats)
    if closegroup_count > 1:
        ws_summury.append([f"Закрытые группы: {closegroup_count-1}"])
        ws_closed_groups = wb.create_sheet("Закрытые группы")
        await write_data(ws_closed_groups, closechats)
    if closegroupdel_count > 1:
        ws_summury.append([f"Удаленные группы: {closegroupdel_count-1}"])
        ws_closed_groups_del = wb.create_sheet("Удаленные группы")
        await write_data_del(ws_closed_groups_del, delgroups)
    if owner_opengroup > 11:
        ws_summury.append([f"Имеет права владельца или админа в открытых группах: {owner_opengroup}"])
    if owner_closegroup > 1:
        ws_summury.append([f"Имеет права владельца или админа в закрытых группах: {owner_closegroup}"])
    
    wb.save(f"{phone}_about.xlsx")

#  Формируем отчет HTML
async def generate_html_report(phone, userid, userinfo, firstname, lastname, username, total_contacts, total_contacts_with_phone, total_mutual_contacts, openchannel_count, closechannel_count,
                               opengroup_count, closegroup_count, closegroupdel_count, owner_openchannel, owner_closechannel, owner_opengroup, owner_closegroup, public_channels_html,
                               private_channels_html, public_groups_html, private_groups_html, deleted_groups_html, blocked_bot_info_html, user_bots_html, user_chat_id,
                               photos_user_html, bot_from_search_html):
    
    # Открываем HTML шаблон
    with open('template.html', 'r', encoding='utf-8') as file:
        template = Template(file.read())

    # Заполняем шаблон данными
    html_content = template.render(
        phone=phone,
        userid=userid,
        firstname=firstname,
        lastname=lastname,
        username=username,
        total_contacts=total_contacts,
        total_contacts_with_phone=total_contacts_with_phone,
        total_mutual_contacts=total_mutual_contacts,
        openchannel_count=openchannel_count,
        closechannel_count=closechannel_count,
        opengroup_count=opengroup_count,
        closegroup_count=closegroup_count,
        closegroupdel_count=closegroupdel_count,
        owner_openchannel=owner_openchannel,
        owner_closechannel=owner_closechannel,
        owner_opengroup=owner_opengroup,
        owner_closegroup=owner_closegroup,
        blocked_bot_info_html=blocked_bot_info_html,
        user_bots_html=user_bots_html,
        public_channels_html=public_channels_html,
        private_channels_html=private_channels_html,
        public_groups_html=public_groups_html,
        private_groups_html=private_groups_html,
        deleted_groups_html=deleted_groups_html,
        user_chat_id=user_chat_id,
        photos_user_html=photos_user_html,
        bot_from_search_html=bot_from_search_html
    )

    # Полный путь до файла отчета на хосте, в монтированной директории
    report_filename = f'/app/files_from_harvester/{phone}_report.html'

    # Записываем результат в HTML файл
    with open(report_filename, 'w', encoding='utf-8') as file:
        file.write(html_content)
